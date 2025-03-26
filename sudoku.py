#A humble attempt to create a Sudoku Solver 
"""
Let's lay down ground rules of understanding the sudoku syntax
We shall follow the matrix rule of using rows X columns, and A-I for boxes
It should look something like this

    1   2    3       4   5   6       7   8   9  
    A11	A12	A13     B14	B15	B16	    C17	C18	C19 1
    A21	A22	A23	    B24	B25	B26	    C27	C28	C29 2
    A31	A32	A33	    B34	B35	B36	    C37	C38	C39 3   
    
    D41	D42	D43	    E44	E45	E46	    F47	F48	F49 4
    D51	D52	D53	    E54	E55	E56	    F57	F58	F59 5   
    D61	D62	D63	    E64	E65	E66	    F67	F68	F69 6
    
    G71	G72	G73	    H74	H75	H76	    I77	I78	I79 7
    G81	G82	G83	    H84	H85	H86	    I87	I88	I89 8   
    G91	G92	G93	    H94	H95	H96	    I97	I98	I99 9


"""
from openpyxl import load_workbook

def box(i,j):
    if 1<=i<=3 and 1<=j<=3:
        return "a"
    elif 1<=i<=3 and 4<=j<=6:
        return "b"
    elif 1<=i<=3 and 6<=j<=9:
        return "c"
    elif 4<=i<=6 and 1<=j<=3:
        return "d"
    elif 4<=i<=6 and 4<=j<=6:
        return "e"
    elif 4<=i<=6 and 6<=j<=9:
        return "f"
    elif 6<=i<=9 and 1<=j<=3:
        return "g"
    elif 6<=i<=9 and 4<=j<=6:
        return "h"
    elif 6<=i<=9 and 6<=j<=9:
        return "i"

def extract_rcb(key):
    # Extract row, column, and box information from the key
    box_init = key[0]  # First character (e.g., 'a')
    i = int(key[1])       # Second character (row number)
    j = int(key[2])       # Third character (column number)

    return box_init,i,j

def update_sudoku(key, value):
    sudoku[key] = value  # Update the value in sudoku


    # Extract row, column, and box information from the key
    box_initial,i,j=extract_rcb(key)

    # Update the related sets
    rows[f"row_{i}"].add(sudoku[key])
    cols[f"col_{j}"].add(sudoku[key])
    boxes[f"box_{box_initial}"].add(sudoku[key])

def pretty_print():
        
    print("-------------------------")
    for i in range(1,10):
        for j in range(1,10):
            box_initial=box(i,j)
            if j%3==0:
                spc=" | "
            else:
                spc=" "
            
            if j==1:
                print("|",end=" ")

            print(sudoku[f"{box_initial}{i}{j}"], end=spc)
        print()
        if i%3==0:
            print("-------------------------")

sudoku={}
rows={}
cols={}
boxes={}
complete={1, 2, 3, 4, 5, 6, 7, 8, 9}
to_be_found=set()


#initialising the sets of rows and columns, ie row_1, col_1 and so on
for i in range(1,10):
    rows[f"row_{i}"]=set()
    cols[f"col_{i}"]=set()
    
for i in "abcdefghi":
    boxes[f"box_{i}"]=set()


for i in range(1,10):
    for j in range(1,10):
        box_initial=box(i,j)
        sudoku[f"{box_initial}{i}{j}"]=0
        #rows[f"row_{i}"].add(sudoku[f"{box_initial}{i}{j}"])
        #cols[f"col_{j}"].add(sudoku[f"{box_initial}{i}{j}"])
        #boxes[f"box_{box_initial}"].add(sudoku[f"{box_initial}{i}{j}"])

#pretty_print()


# Load the workbook
workbook = load_workbook("sudoku_input.xlsx")  # Replace with your file path

# Select a worksheet
sheet = workbook.active  # Or workbook["SheetName"]
values=[]
for row in sheet.iter_rows(values_only=True):
    #print(type(row))
    row=list(row)
    values.extend(row)

#loading the values in the sudoku dict
flag=0
for yo in sudoku:
    #print(yo)
    sudoku[yo]=values[flag]
    update_sudoku(yo,values[flag])
    flag+=1

#pretty_print()
#x=complete-rows["row_1"]-cols["col_1"]-boxes["box_a"]

for key in sudoku:
    #print(key,sudoku[key])
    if sudoku[key]==None:
        to_be_found.add(key)

found_already=set()

print(f"Length of to_be_found is {len(to_be_found)}")
yo=0
while yo<=5:
    for unknown in to_be_found:
        box_init,i,j=extract_rcb(unknown)
        x=complete-rows[f"row_{i}"]-cols[f"col_{j}"]-boxes[f"box_{box_init}"]-{0}-{None}
        #print(x)
        if len(x)==1 or len(x)==2:
            x=list(x)
            update_sudoku(f"{box_init}{i}{j}",x)
            #update_sudoku(f"{box_init}{i}{j}",x[0])
            found_already.add(unknown)
    to_be_found=to_be_found-found_already
    
    print(f"yo= {yo}")
    print(f"Length of to_be_found after iteration is {len(to_be_found)}")
    pretty_print()

    if len(to_be_found)==0:
        break

    if yo>=5:
        dec=input("Do you wanna continue the loop for a few more times, y or n: ")
        if dec=='y':
            yo=0

    
    yo+=1

pretty_print()
    
#print(complete)
#print(boxes["box_a"])


        

#print(sudoku)
