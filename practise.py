#practise for sudoku 
"""
a={1,2}
a=list(a)
print(a[0])

for i in range(1,10):
    for j in range(1,10):
        print(f"{i}{j}", end=' ')

"""

from openpyxl import load_workbook


# Load the workbook
workbook = load_workbook("sudoku_input.xlsx")  # Replace with your file path

# Select a worksheet
sheet = workbook.active  # Or workbook["SheetName"]
values=[]
for row in sheet.iter_rows(values_only=True):
    #print(type(row))
    row=list(row)
    values.extend(row)

print(values)


